import time
import re
import os.path
from datetime import datetime
from PySimpleGUI import popup_get_file
from openpyxl import workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from Ua import ua
"""
加入地址必须要有门牌号
"""


class Checklist(object):

    def __init__(self):
        self.options = Options()
        # self.options.add_argument('--start-maximized')
        self.options.add_argument('--no-sandbox')
        self.options.add_argument("--disable-blink-features=AutomationControlled")
        self.options.add_argument('--disable-javascript')
        self.options.add_experimental_option('useAutomationExtension', False)
        self.options.add_experimental_option('excludeSwitches', ['enable-automation'])
        self.driver = webdriver.Chrome(executable_path="./chromedriver.exe", options=self.options)
        self.wait = WebDriverWait(self.driver, 3)
        self.get_url = 'https://www.google.com/maps/'
        self.driver.get(self.get_url)
        self.wb = workbook.Workbook()
        self.ws = self.wb.active
        # 用于文件名
        self.date = datetime.now().strftime("%Y%m%d")
        self.file_name = ""

    def find_func(self, address, country):
        """
        封装在谷歌查找地址的方法
        :param address: 拼接的地址
        :param country: 用于判断谷歌当前查询的国家是否和表格的国家对应
        :return:
        """
        self.options.add_argument('user-agent={}'.format(ua()))
        print(f'address：{address}, country：{country}')
        # 捕获用于在没查到地址时
        try:
            self.wait.until(
                EC.presence_of_element_located((By.ID, "searchboxinput"))).send_keys(address)
            self.wait.until(
                EC.element_to_be_clickable((By.ID, 'searchbox-searchbutton'))).click()
            time.sleep(4)
            # 一次为获取邮编
            post_code = self.wait.until(
                EC.presence_of_element_located((By.XPATH, '//div[@class="x3AX1-LfntMc-header-title-ij8cu"]'))
            )
            post_code_h2 = post_code.find_elements_by_tag_name('h2')
            item = []
            for h2 in post_code_h2:
                # 获取当前h2标签文本添加到列表
                item.append(h2.text)
            # 判断国家是否存在
            if country not in item:
                return "No"
            h2_splicing = ''.join(item)
            pattern = re.compile(r'\d+')
            extract = pattern.findall(h2_splicing)
            h2_end = ''.join(extract)
            return h2_end
        except TimeoutException:
            return "No"
        finally:
            self.wait.until(
                EC.element_to_be_clickable((By.ID, 'searchboxinput'))).clear()

    @staticmethod
    def excel_read(file_path):
        """
        # 读取表格数据
        :param file_path:
        :return:
        """
        wb = load_workbook(file_path)
        ws = wb.active
        rows = []
        for row in ws.iter_rows():
            rows.append(row)
        for x in range(1, len(rows)):
            data_text = []
            order = str(rows[x][1].value)
            city = str(rows[x][10].value)
            mail_address_first = str(rows[x][11].value)
            mail_address_second = str(rows[x][12].value)
            country = str(rows[x][8].value)
            data_text.append(order)
            data_text.append(city)
            data_text.append(mail_address_first)
            data_text.append(mail_address_second)
            data_text.append(country)
            print(data_text)
            # data_text = ["订单编号", "所属城市", "邮寄地址1(完整导出)", "邮寄地址2", "国家(中)"]
            yield data_text

    @staticmethod
    def address_join(start, end, data_excel):
        address_result = ' '.join(data_excel[start:end])
        return address_result

    @property
    def file_names(self):
        return os.path.basename(self.file_name).strip('.xlsx')

    # 存储没有匹配到邮编对应的订单号
    def excel_save(self, order, i):
        self.ws[f'A{i}'] = order
        file_location = f'./{self.file_names}{self.date}.xlsx'
        self.wb.save(file_location)
        print(f'{order} 存储地址为：{file_location}')

    @property
    def gui_choose_file(self):
        file_address = popup_get_file('请选择你要读取的表格文件：')
        self.file_name = file_address
        return file_address

    # 判断地址是否存在数字，用于区别地址是否有门牌号
    @staticmethod
    def is_number(number):
        pattern = re.compile('[0-9]+')
        match = pattern.findall(number)
        if match:
            return True
        else:
            return False

    def run(self):
        count = 1
        for data_excel in self.excel_read(self.gui_choose_file):
            # 查找地址3列
            address_1_4 = self.address_join(1, 4, data_excel)
            if self.is_number(address_1_4):
                # 邮编长度大于4则视为已查到
                if len(self.find_func(address_1_4, data_excel[4])) >= 4:
                    print(f'{data_excel[0]} 1-4已查到地址')
                    continue

            # 查找地址前两列
            address_1_3 = self.address_join(1, 3, data_excel)
            if self.is_number(address_1_3):
                if len(self.find_func(address_1_3, data_excel[4])) >= 4:
                    print(f'{data_excel[0]} 1-3已查到地址')
                    continue

            # 查找地址后两列
            address_2_4 = self.address_join(2, 4, data_excel)
            if self.is_number(address_2_4):
                if len(self.find_func(address_2_4, data_excel[4])) >= 4:
                    print(f'{data_excel[0]} 2-4已查到地址')
                    continue

            self.excel_save(data_excel[0], count)
            count += 1
        print("查询完毕")


if __name__ == '__main__':
    ds = Checklist()
    ds.run()

